
from django.shortcuts import render

# Create your views here.
import csv
from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.viewsets import ModelViewSet
from rest_framework.parsers import MultiPartParser, FormParser

from apps.client_customer.models import ClientCustomerDependent
from apps.master_management.models import MasterRelationship
from .models import CareProgram , EyeDentalTreatment , MedicalCamp , CampCase , CHP , TypeOfOHC , OHC , EyeTreatmentCase , DentalTreatmentCase
from apps.test_individual.models import IndividualTest as Test
from .serializers import CareProgramSerializer , EyeDentalTreatmentSerializer , MedicalCampSerializer , CampCaseSerializer , CHPSerializer, TypeOfOHCSerializer , OHCSerializer , EyeTreatmentCaseSerializer , DentalTreatmentCaseSerializer
from rest_framework.permissions import IsAdminUser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework.response import Response as response

from rest_framework import status


class CareProgramViewSet(ModelViewSet):
    queryset = CareProgram.objects.all().order_by('-id')
    serializer_class = CareProgramSerializer
    parser_classes = (MultiPartParser, FormParser)
    permission_classes=[IsAdminUser]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['is_active']          # filter by active/inactive
    search_fields = ['care_program_name'] 


    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="care_programs.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID',
            'Care Program Name',
            'Normal Price',
            'Care Program Details',
            'Is Active',
            'Created At'
        ])

        for obj in queryset:
            writer.writerow([
                obj.id,
                obj.care_program_name,
                obj.normal_price,
                obj.care_program_details,
                obj.is_active,
                obj.created_at
            ])

        return response
    

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Care Programs"

    # Header row
        headers = [
            "ID",
            "Care Program Name",
            "Normal Price",
            "Care Program Details",
            "Is Active",
            "Created At"
        ]
        ws.append(headers)

    # Data rows
        for obj in queryset:
            ws.append([
                obj.id,
                obj.care_program_name,
                obj.normal_price,
                obj.care_program_details,
                obj.is_active,
                obj.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

    # Auto-adjust column width
        for col_num, column_title in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 22

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=care_programs.xlsx'

        wb.save(response)
        return response


# EYE & DENTAL TREATMENTS VIEWSET-----


class EyeDentalTreatmentViewSet(ModelViewSet):
    queryset = EyeDentalTreatment.objects.all().order_by("-id")
    serializer_class = EyeDentalTreatmentSerializer
    permission_classes = [IsAdminUser]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def get_queryset(self):
        qs = super().get_queryset()

        treatment_type = self.request.query_params.get("type")
        treatment_name = self.request.query_params.get("treatment_name")
        # is_active = self.request.query_params.get("is_active")

        if treatment_type:
            qs = qs.filter(treatment_type__iexact=treatment_type)

        if treatment_name:
            qs = qs.filter(treatment_name__icontains=treatment_name)

        # if is_active is not None:
        #     if is_active.lower() in ["true", "1", "yes"]:
        #         qs = qs.filter(is_active=True)
        #     elif is_active.lower() in ["false", "0", "no"]:
        #         qs = qs.filter(is_active=False)

        return qs
    


# MEDICAL CAMP VIEWSET-----


class MedicalCampViewSet(ModelViewSet):
    queryset = MedicalCamp.objects.all().order_by("-id")
    serializer_class = MedicalCampSerializer
    permission_classes = [IsAdminUser]
    queryset = MedicalCamp.objects.select_related(
        "main_client", "sub_client", "package", "tests", "camp_state", "camp_city", "network_provider"
    )

    
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):

        serializer.save(updated_by=self.request.user)
        instance = self.get_object()

        # Get the new value user is sending (or keep old if not provided)
        new_request_from = self.request.data.get("camp_request_from", instance.camp_request_from)

        camp = serializer.save()

        # ✅ If changed to "Main Client", clear sub_client
        if new_request_from == "Main Client":
            if camp.sub_client is not None:
                camp.sub_client = None
                camp.save(update_fields=["sub_client"])

    # FILTERS-----

    def apply_filters(self, qs, request):
        camp_request_from = request.query_params.get("camp_request_from")
        client_name = request.query_params.get("client_name")
        sub_client = request.query_params.get("sub_client")
        camp_status = request.query_params.get("camp_status")

        if camp_request_from:
            qs = qs.filter(camp_request_from=camp_request_from)

        if client_name:
            qs = qs.filter(main_client__corporate_name__icontains=client_name)

        if sub_client:
            qs = qs.filter(sub_client_id=sub_client)

        if camp_status:
            qs = qs.filter(camp_status=camp_status)

        return qs
    
    # CAMP LIST---
    @action(detail=False, methods=["get"], url_path="camp-list")
    def camp_list(self, request):
        qs = MedicalCamp.objects.exclude(camp_status="Completed")

        camp_request_from = request.query_params.get("camp_request_from")
        main_client = request.query_params.get("main_client")  
        sub_client = request.query_params.get("sub_client")   
        camp_status = request.query_params.get("camp_status") 

        if camp_request_from:
            values = [v.strip() for v in camp_request_from.split(",") if v.strip()]
            qs = qs.filter(camp_request_from__in=values)

        if main_client:
            ids = [i.strip() for i in main_client.split(",") if i.strip()]
            qs = qs.filter(main_client_id__in=ids)

        if sub_client:
            ids = [i.strip() for i in sub_client.split(",") if i.strip()]
            qs = qs.filter(sub_client_id__in=ids)

        # ✅ Multiple statuses: camp_status=Fresh Case,On Going
        if camp_status:
            statuses = [s.strip() for s in camp_status.split(",") if s.strip()]
            qs = qs.filter(camp_status__in=statuses)

        serializer = MedicalCampSerializer(qs, many=True)
        return response(serializer.data)
    

    # CLOSED/CAMP COMPLETED LIST---
    @action(detail=False, methods=["get"], url_path="closed-list")
    def closed_list(self, request):
        qs = MedicalCamp.objects.filter(camp_status="Completed")

        camp_request_from = request.query_params.get("camp_request_from")
        main_client = request.query_params.get("main_client")
        sub_client = request.query_params.get("sub_client")
        camp_status = request.query_params.get("camp_status")

        if camp_request_from:
            values = [v.strip() for v in camp_request_from.split(",") if v.strip()]
            qs = qs.filter(camp_request_from__in=values)

        if main_client:
            ids = [i.strip() for i in main_client.split(",") if i.strip()]
            qs = qs.filter(main_client_id__in=ids)

        if sub_client:
            ids = [i.strip() for i in sub_client.split(",") if i.strip()]
            qs = qs.filter(sub_client_id__in=ids)

        if camp_status:
            statuses = [s.strip() for s in camp_status.split(",") if s.strip()]
            qs = qs.filter(camp_status__in=statuses)

        serializer = MedicalCampSerializer(qs, many=True)
        return response(serializer.data)
        

# CASE VIEWSET-----

class CampCaseViewSet(ModelViewSet):
    queryset = CampCase.objects.select_related(
        "camp",
        "camp__main_client",
        
    )
    serializer_class = CampCaseSerializer
    permission_classes = [IsAdminUser]


    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


    # CASE LIST WITH FILTERS-----
    def get_queryset(self):
        qs = super().get_queryset()

        # Query params
        camp_id = self.request.query_params.get("camp_id")        # MedicalCamp ID
        client_name = self.request.query_params.get("client_name")  # text

        # Filter by camp (FK id)
        if camp_id:
            ids = [i.strip() for i in camp_id.split(",") if i.strip()]
            qs = qs.filter(camp_id__in=ids)
        # Filter by client name (from MedicalCamp -> Client)
        if client_name:
            names = [n.strip() for n in client_name.split(",") if n.strip()]
            qs = qs.filter(camp__main_client__corporate_name__in=names)

        return qs
    

# CHP VIEWSET-----


class CHPViewSet(ModelViewSet):
    queryset = CHP.objects.select_related("package", "product" , "service")
    serializer_class = CHPSerializer
    permission_classes = [IsAdminUser]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    # FLITERS---

    
    def get_queryset(self):
        qs = super().get_queryset()

        
        package = self.request.query_params.get("package")     
        product = self.request.query_params.get("product")     
        service = self.request.query_params.get("service")   
        category = self.request.query_params.get("category")    
        frequency = self.request.query_params.get("frequency")  

        if package:
            ids = [i.strip() for i in package.split(",") if i.strip()]
            qs = qs.filter(package_id__in=ids)

        if product:
            ids = [i.strip() for i in product.split(",") if i.strip()]
            qs = qs.filter(product_id__in=ids)

        if service:
            ids = [i.strip() for i in service.split(",") if i.strip()]
            qs = qs.filter(service_id__in=ids)

        if category:
            values = [v.strip() for v in category.split(",") if v.strip()]
            qs = qs.filter(category__in=values)

        if frequency:
            values = [v.strip() for v in frequency.split(",") if v.strip()]
            qs = qs.filter(frequency__in=values)

        return qs
    

# OHC  MASTER VIEWSET-----

class TypeOfOHCViewSet(ModelViewSet):
    queryset = TypeOfOHC.objects.all()
    serializer_class = TypeOfOHCSerializer
    permission_classes = [IsAdminUser]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


# OHC MAIN VIEWSET-----

class OHCViewSet(ModelViewSet):
    queryset = OHC.objects.select_related("type_of_ohc", "client", "doctor")
    serializer_class = OHCSerializer
    permission_classes = [IsAdminUser]


    def _get_doctor_qualification_text(self, doctor):
       
        qs = doctor.qualifications.all()
        names = [q.name for q in qs]  
        return ", ".join(names)


    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        doctor = data.get("doctor")

        # ✅ Auto-fetch qualification from Doctor
        qualification = self._get_doctor_qualification_text(doctor)

        ohc = OHC.objects.create(
            type_of_ohc=data.get("type_of_ohc"),
            client=data.get("client"),
            doctor=doctor,

            corporate_requirements=data.get("corporate_requirements", ""),
            crm_name=data.get("crm_name", ""),
            corporate_address=data.get("corporate_address", ""),

            spoc_name=data.get("spoc_name", ""),
            spoc_email=data.get("spoc_email", ""),
            spoc_mobile=data.get("spoc_mobile", ""),

            service_start_date=data.get("service_start_date"),
            agreement_date=data.get("agreement_date"),
            relationship_end_date=data.get("relationship_end_date"),

            agreement_upload=data.get("agreement_upload"),

            client_bill_amount=data.get("client_bill_amount"),
            service_provider_cost=data.get("service_provider_cost"),

            doctor_qualifications=qualification,                 # 🔥 auto-filled
            doctor_certificate_link=data.get("doctor_certificate_link", ""),  # from frontend

            remarks=data.get("remarks", ""),
            created_by=request.user,
            updated_by=request.user,
        )

        out = self.get_serializer(ohc)
        return response(out.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        doctor = data.get("doctor", instance.doctor)

        # If doctor changed, refresh qualification
        qualification = self._get_doctor_qualification_text(doctor)

        # Update fields
        for attr, value in data.items():
            setattr(instance, attr, value)

        instance.doctor = doctor
        instance.doctor_qualifications = qualification  # 🔥 always sync from Doctor
        instance.updated_by = request.user
        instance.save()

        out = self.get_serializer(instance)
        return response(out.data, status=status.HTTP_200_OK)


    def get_queryset(self):
        qs = super().get_queryset()

        type_of_ohc = self.request.query_params.get("type_of_ohc")  
        client = self.request.query_params.get("client")            
        doctor = self.request.query_params.get("doctor")           

        if type_of_ohc:
            ids = [i.strip() for i in type_of_ohc.split(",") if i.strip()]
            qs = qs.filter(type_of_ohc_id__in=ids)

        if client:
            ids = [i.strip() for i in client.split(",") if i.strip()]
            qs = qs.filter(client_id__in=ids)

        if doctor:
            ids = [i.strip() for i in doctor.split(",") if i.strip()]
            qs = qs.filter(doctor_id__in=ids)

        return qs
    

# EYE PROCEDURE CASE VIEWSET-----


class EyeTreatmentCaseViewSet(ModelViewSet):
    queryset = EyeTreatmentCase.objects.all().order_by('-id')
    serializer_class = EyeTreatmentCaseSerializer
    permission_classes = [IsAdminUser]



    @action(detail=False, methods=['get'], url_path='relationship-persons')
    def relationship_persons(self, request):
        employee_id = request.query_params.get('employee_id')
        case_for_id = request.query_params.get('case_for_id')

        relationship = MasterRelationship.objects.get(id=case_for_id)

        # SELF → no dropdown
        if relationship.name.lower() == 'self':
            return response([])

        dependants = ClientCustomerDependent.objects.filter(
            customer_id=employee_id,
            relationship_id=case_for_id
        )

        return response([
            {
                "id": d.id,
                "name": d.name
            } for d in dependants
        ])
    

    @action(detail=False, methods=['get'], url_path='eye-treatments')
    def eye_treatments(self, request):
        qs = EyeDentalTreatment.objects.filter(
            treatment_type='Eye',
            is_active=True
        ).order_by('name')

        return response([
            {"id": t.id, "name": t.name}
            for t in qs
        ])

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


# DENTAL PROCEDURE CASE VIEWSET-----



class DentalTreatmentCaseViewSet(ModelViewSet):
    queryset = DentalTreatmentCase.objects.all().order_by('-id')
    serializer_class = DentalTreatmentCaseSerializer
    permission_classes = [IsAdminUser]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['get'], url_path='relationship-persons')
    def relationship_persons(self, request):
        employee_id = request.query_params.get('employee_id')
        case_for_id = request.query_params.get('case_for_id')

        if not employee_id or not case_for_id:
            return response([])

        relationship = MasterRelationship.objects.get(id=case_for_id)
        if relationship.name.lower() == 'self':
            return response([])

        dependants = ClientCustomerDependent.objects.filter(
            customer_id=employee_id,
            relationship_id=case_for_id
        )

        return response([{"id": d.id, "name": d.name} for d in dependants])

    @action(detail=False, methods=['get'], url_path='dental-treatments')
    def dental_treatments(self, request):
        qs = EyeDentalTreatment.objects.filter(
            treatment_type='Dental',
            is_active=True
        ).order_by('name')

        return response([{"id": t.id, "name": t.name} for t in qs])


   